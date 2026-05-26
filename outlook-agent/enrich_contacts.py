"""
enrich_contacts.py
==================
Phase 2 of the Outlook agent.

Reads external_contacts.xlsx, identifies rows with missing phone / position /
company, scans every matching email thread in Outlook, parses the sender's
signature block, and writes the enriched data back to the same Excel file.

Run after outlook-agent.py:
    python enrich_contacts.py
"""

import re
import html
import sys
import os
from collections import defaultdict

try:
    import win32com.client
except ImportError:
    sys.exit("ERROR: pip install pywin32")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    sys.exit("ERROR: pip install openpyxl")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  CONFIG
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
EXCEL_PATH = r'C:\Users\maxbi\OneDrive\Documents\GitHub\My-Agents\outlook-agent\external_contacts.xlsx'

# Columns in the Excel (1-based)
COL_NAME     = 1
COL_COMPANY  = 2
COL_POSITION = 3
COL_EMAIL    = 4
COL_PHONE    = 5

# Max email bodies to examine per contact (keeps runtime reasonable)
MAX_BODIES = 20

# Outlook item class
OL_MAIL_ITEM = 43

SKIP_FOLDERS = {
    'deleted items', 'recoverable items', 'drafts', 'outbox',
    'junk email', 'spam', 'trash', 'clutter',
    'conversation history', 'rss feeds', 'sync issues',
    'conflicts', 'local failures', 'server failures',
    'calendar', 'tasks', 'notes', 'contacts',
    'recipient cache', 'personmetadata',
}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  REGEX PATTERNS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Phone: Israeli (local & +972) + generic international
PHONE_RE = re.compile(
    r'(?:(?:mobile|mob|cell|tel|phone|direct|fax|m|t|p)\s*[:\-\.]?\s*)?'
    r'(?:'
        r'\+972[\s\-\.]?(?:0)?[2-9]\d[\s\-\.]?\d{3}[\s\-\.]?\d{4}'   # +972-XX-XXXXXXX
        r'|0[2-9]\d[\s\-\.]?\d{3}[\s\-\.]?\d{4}'                       # 0XX-XXXXXXX
        r'|\+[1-9]\d{0,2}[\s\-\.]?\(?\d{1,4}\)?[\s\-\.]?\d{2,4}[\s\-\.]?\d{2,4}(?:[\s\-\.]?\d{0,4})?'  # +CC ...
    r')',
    re.IGNORECASE,
)

# Lines that look like job titles
TITLE_RE = re.compile(
    r'\b(?:'
    r'ceo|cto|coo|cfo|cso|cpo|vp\b|vice[\s\-]president|president|'
    r'director|manager|engineer|engineering|developer|analyst|'
    r'consultant|founder|co[\-\s]?founder|partner|'
    r'head\s+of|team\s+lead|tech\s+lead|lead\b|'
    r'senior\b|sr\b|junior\b|jr\b|principal|'
    r'account\s+manager|key\s+account|sales|'
    r'business\s+development|bd\b|'
    r'procurement|purchasing|supply\s+chain|sourcing|logistics|'
    r'field\s+application|application\s+engineer|fae\b|'
    r'hardware|software|electronics|embedded|'
    r'product\s+manager|project\s+manager|program\s+manager|'
    r'operations|quality|reliability|compliance|'
    r'regional|country|national|global|'
    r'r\s*&\s*d|research|technical\b|technician'
    r')',
    re.IGNORECASE,
)

# Company suffixes that confirm a line is a company name
COMPANY_SUFFIX_RE = re.compile(
    r'\b(?:ltd\.?|limited|inc\.?|incorporated|corp\.?|corporation|'
    r'co\.(?:\s|$)|llc|l\.l\.c|gmbh|b\.v\.|bv\b|a\.g\.|ag\b|'
    r'plc|s\.a\.|s\.r\.l\.?|pty|oy|as\b|group|industries|'
    r'systems|solutions|technologies|electronics|engineering)'
    r'|בע"מ|בעמ',
    re.IGNORECASE,
)

# Signature delimiter patterns (marks the start of the signature block)
SIG_DELIM_RE = re.compile(
    r'^(?:'
    r'--|_{2,}|={2,}|\*{2,}'          # -- or ___ or === or ***
    r'|(?:best\s*)?regards?'
    r'|kind\s+regards?'
    r'|warm\s+regards?'
    r'|with\s+(?:best\s+)?regards?'
    r'|sincerely'
    r'|yours\s+(?:truly|sincerely|faithfully)'
    r'|thank(?:s|\s+you)'
    r'|cheers'
    r'|all\s+the\s+best'
    r'|have\s+a\s+(?:great|good|nice)'
    r'|בברכה|בכבוד|תודה|שלום'
    r')\s*[,.]?\s*$',
    re.IGNORECASE,
)

# Lines to discard from signature (quoted mail headers, URLs, disclaimers)
DISCARD_LINE_RE = re.compile(
    r'(?:from:|to:|sent:|cc:|subject:|date:|on\s+\w+.*wrote:|'
    r'http[s]?://|www\.|@\S+\.\S+|'
    r'confidential|disclaimer|this\s+email|this\s+message|'
    r'unsubscribe|privacy\s+policy)',
    re.IGNORECASE,
)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  HTML → PLAIN TEXT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def html_to_text(raw: str) -> str:
    text = re.sub(r'<(?:br|BR)\s*/?>', '\n', raw)
    text = re.sub(r'</?(?:p|P|div|DIV|tr|TR|li|LI|h\d|H\d)[^>]*>', '\n', text)
    text = re.sub(r'<[^>]+>', '', text)
    text = html.unescape(text)
    return text


def get_plain_body(mail) -> str:
    try:
        body = getattr(mail, 'Body', '') or ''
        if body.strip():
            return body
        html_body = getattr(mail, 'HTMLBody', '') or ''
        if html_body.strip():
            return html_to_text(html_body)
    except Exception:
        pass
    return ''


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SIGNATURE EXTRACTION
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def extract_signature_lines(body: str) -> list[str]:
    """
    Return the lines most likely to be the sender's signature block.
    Strategy: find a signature delimiter or a closing phrase, then take
    the following lines (max 25). If nothing found, take the last 20 lines.
    Stop at quoted-reply markers.
    """
    lines = body.splitlines()

    # Find the first quoted-reply boundary (>From, "On ... wrote:", "-----Original")
    reply_boundary = len(lines)
    for i, line in enumerate(lines):
        stripped = line.strip()
        if re.match(r'^(?:>+\s*From|-----\s*Original\s*Message|_{5,})', stripped, re.IGNORECASE):
            reply_boundary = i
            break

    lines = lines[:reply_boundary]

    # Find signature delimiter
    for i, line in enumerate(lines):
        if SIG_DELIM_RE.match(line.strip()):
            return [l for l in lines[i: i + 25] if not DISCARD_LINE_RE.search(l)]

    # Fall back: last 20 lines
    return [l for l in lines[-20:] if not DISCARD_LINE_RE.search(l)]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  FIELD PARSERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def parse_phone(lines: list[str]) -> str:
    for line in lines:
        m = PHONE_RE.search(line)
        if m:
            raw = m.group().strip()
            # Clean up separators, keep digits + + - space
            cleaned = re.sub(r'[^\d\+\-\s\(\)]', '', raw).strip()
            if len(re.sub(r'\D', '', cleaned)) >= 7:
                return cleaned
    return ''


def parse_position(lines: list[str], known_name: str = '') -> str:
    name_lower = known_name.strip().lower()
    candidates = []
    for line in lines:
        stripped = line.strip()
        if not stripped or len(stripped) > 80:
            continue
        # Skip lines that are just the person's name
        if name_lower and stripped.lower() == name_lower:
            continue
        # Skip lines that look like phone numbers or email addresses
        if PHONE_RE.search(stripped) or '@' in stripped or re.match(r'^https?://', stripped):
            continue
        if TITLE_RE.search(stripped):
            candidates.append(stripped)

    # Prefer shorter, cleaner candidates
    candidates.sort(key=lambda x: len(x))
    return candidates[0] if candidates else ''


def parse_company(lines: list[str], email_domain: str = '',
                  known_position: str = '') -> str:
    # 1. Look for a line with explicit company suffix
    for line in lines:
        stripped = line.strip()
        if not stripped or len(stripped) > 80:
            continue
        if '@' in stripped or PHONE_RE.search(stripped):
            continue
        if COMPANY_SUFFIX_RE.search(stripped):
            return stripped

    # 2. Fall back: infer from domain
    if email_domain:
        domain = email_domain.split('.')[0]
        if domain not in ('gmail', 'yahoo', 'hotmail', 'outlook', 'icloud', 'me'):
            return domain.replace('-', ' ').replace('_', ' ').title()

    return ''


def parse_signature(sig_lines: list[str], known_name: str = '',
                    email_domain: str = '') -> dict:
    phone    = parse_phone(sig_lines)
    position = parse_position(sig_lines, known_name)
    company  = parse_company(sig_lines, email_domain, position)
    return {'phone': phone, 'position': position, 'company': company}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SCORING — pick the best data seen across multiple email bodies
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def score(d: dict) -> int:
    return bool(d.get('phone')) + bool(d.get('position')) + bool(d.get('company'))


def merge_best(acc: dict, new: dict) -> dict:
    """Keep the field value from whichever record is richer."""
    result = dict(acc)
    for field in ('phone', 'position', 'company'):
        if not result.get(field) and new.get(field):
            result[field] = new[field]
    return result


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  EXCEL LOADER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def load_excel(path: str):
    """
    Returns:
        wb          — workbook object
        ws          — active worksheet
        targets     — dict: email -> {'row': int, 'name': str,
                                       'needs': set of missing field names}
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    targets = {}

    for row in ws.iter_rows(min_row=2):
        email    = (row[COL_EMAIL    - 1].value or '').strip().lower()
        name     = (row[COL_NAME     - 1].value or '').strip()
        company  = (row[COL_COMPANY  - 1].value or '').strip()
        position = (row[COL_POSITION - 1].value or '').strip()
        phone    = (row[COL_PHONE    - 1].value or '').strip()

        if not email:
            continue

        needs = set()
        if not phone:    needs.add('phone')
        if not position: needs.add('position')
        if not company:  needs.add('company')

        if needs:
            targets[email] = {
                'row':      row[0].row,
                'name':     name,
                'needs':    needs,
                'found':    {},  # will be filled during scan
            }

    return wb, ws, targets


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  OUTLOOK SENDER SMTP (reused from main agent)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
_EMAIL_RE = re.compile(r'[\w.+\-]+@[\w\-]+\.[\w.\-]+')

def extract_smtp(raw: str) -> str:
    if not raw:
        return ''
    m = re.search(r'<([^>]+)>', raw)
    if m and '@' in m.group(1):
        return m.group(1).strip().lower()
    m = _EMAIL_RE.search(raw)
    return m.group().lower() if m else ''


def get_sender_smtp(mail) -> str:
    try:
        if getattr(mail, 'SenderEmailType', '') == 'EX':
            try:
                exch = mail.Sender.GetExchangeUser()
                if exch:
                    return extract_smtp(exch.PrimarySmtpAddress)
            except Exception:
                pass
        addr = getattr(mail, 'SenderEmailAddress', '') or ''
        if '@' in addr:
            return extract_smtp(addr)
    except Exception:
        pass
    return ''


def get_recipients_smtp(mail) -> list:
    results = []
    try:
        for i in range(1, mail.Recipients.Count + 1):
            try:
                r = mail.Recipients.Item(i)
                addr = getattr(r, 'Address', '')
                if addr and '@' in addr and not addr.startswith('/O='):
                    results.append(extract_smtp(addr))
                else:
                    entry = getattr(r, 'AddressEntry', None)
                    if entry:
                        try:
                            exch = entry.GetExchangeUser()
                            if exch:
                                results.append(extract_smtp(exch.PrimarySmtpAddress))
                        except Exception:
                            pass
            except Exception:
                pass
    except Exception:
        pass
    return [r for r in results if r]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  FOLDER WALKER — collect bodies for target contacts
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def walk_and_collect(folder, targets: dict, body_pool: dict,
                     is_sent_tree=False, depth=0, max_depth=12):
    if depth > max_depth:
        return
    try:
        name_lower = folder.Name.strip().lower()
    except Exception:
        return
    if name_lower in SKIP_FOLDERS:
        return

    is_sent = is_sent_tree or name_lower in ('sent items', 'sent mail', 'sent')

    try:
        items = folder.Items
        count = items.Count
        if count:
            print(f"    [{folder.Name}] {count}", end=' ... ', flush=True)
            found = 0
            for i in range(1, count + 1):
                try:
                    item = items.Item(i)
                    if getattr(item, 'Class', 0) != OL_MAIL_ITEM:
                        continue

                    if is_sent:
                        # Sent mail: check if any recipient is a target
                        for addr in get_recipients_smtp(item):
                            if addr in targets and len(body_pool.get(addr, [])) < MAX_BODIES:
                                body = get_plain_body(item)
                                if body.strip():
                                    body_pool.setdefault(addr, []).append(body)
                                    found += 1
                    else:
                        # Received mail: check sender
                        sender = get_sender_smtp(item)
                        if sender in targets and len(body_pool.get(sender, [])) < MAX_BODIES:
                            body = get_plain_body(item)
                            if body.strip():
                                body_pool.setdefault(sender, []).append(body)
                                found += 1
                except Exception:
                    pass
            print(f"{found} matched")
    except Exception as e:
        print(f"    [warn] {folder.Name}: {e}")

    try:
        for j in range(1, folder.Folders.Count + 1):
            try:
                walk_and_collect(
                    folder.Folders.Item(j), targets, body_pool,
                    is_sent_tree=is_sent, depth=depth + 1, max_depth=max_depth,
                )
            except Exception:
                pass
    except Exception:
        pass


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  EXCEL WRITER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
_UPDATED_FILL = PatternFill('solid', fgColor='D6F0D6')  # light green = newly filled
_UPDATED_FONT = Font(name='Calibri', size=10, color='1A5C1A')


def update_excel(wb, ws, targets: dict, path: str) -> int:
    updated_cells = 0
    for email, info in targets.items():
        found  = info.get('found', {})
        row    = info['row']
        if not found:
            continue

        field_col = {
            'company':  COL_COMPANY,
            'position': COL_POSITION,
            'phone':    COL_PHONE,
        }
        for field, col in field_col.items():
            val = found.get(field, '')
            if val and field in info['needs']:
                cell = ws.cell(row=row, column=col)
                if not (cell.value or '').strip():
                    cell.value = val
                    cell.fill  = _UPDATED_FILL
                    cell.font  = _UPDATED_FONT
                    updated_cells += 1

    wb.save(path)
    return updated_cells


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  MAIN
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def main():
    print("=" * 60)
    print("  Outlook Contacts Enrichment (Phase 2)")
    print("=" * 60)

    # ── Load Excel ────────────────────────────────────────────────
    if not os.path.exists(EXCEL_PATH):
        sys.exit(f"ERROR: Excel not found:\n  {EXCEL_PATH}\nRun outlook-agent.py first.")

    print(f"\nLoading Excel: {EXCEL_PATH}")
    wb, ws, targets = load_excel(EXCEL_PATH)
    print(f"  {ws.max_row - 1} total contacts")
    print(f"  {len(targets)} need enrichment (missing phone / position / company)")

    if not targets:
        print("\nAll contacts are already fully populated. Nothing to do.")
        return

    # ── Connect to Outlook ────────────────────────────────────────
    print("\nConnecting to Outlook...")
    try:
        outlook = win32com.client.Dispatch('Outlook.Application')
        ns      = outlook.GetNamespace('MAPI')
    except Exception as e:
        sys.exit(f"ERROR: {e}")

    # ── Walk all stores and collect bodies ────────────────────────
    body_pool: dict[str, list[str]] = {}

    store_count = ns.Stores.Count
    print(f"\nScanning {store_count} store(s) for email threads...")
    for s in range(1, store_count + 1):
        try:
            store = ns.Stores.Item(s)
            sname = getattr(store, 'DisplayName', f'Store {s}')
            print(f"\n-- {sname} --")
            walk_and_collect(store.GetRootFolder(), targets, body_pool)
        except Exception as e:
            print(f"  [warn] store {s}: {e}")

    print(f"\nEmail bodies collected for {len(body_pool)} contact(s).")

    # ── Parse signatures ──────────────────────────────────────────
    print("\nParsing signatures...")
    enriched_count = 0

    for email, bodies in body_pool.items():
        info   = targets.get(email)
        if not info:
            continue

        domain  = email.split('@')[-1] if '@' in email else ''
        name    = info.get('name', '')
        best    = {}

        for body in bodies:
            sig_lines = extract_signature_lines(body)
            parsed    = parse_signature(sig_lines, known_name=name, email_domain=domain)

            # Only use parsed fields that address missing data
            filtered = {k: v for k, v in parsed.items()
                        if k in info['needs'] and v}
            best = merge_best(best, filtered)

            if score(best) == len(info['needs']):
                break  # found everything needed

        if best:
            targets[email]['found'] = best
            enriched_count += 1
            fields = ', '.join(f"{k}={v}" for k, v in best.items() if v)
            print(f"  [{email}] {fields}".encode('ascii', errors='replace').decode('ascii'))

    print(f"\n{enriched_count} contact(s) enriched from signatures.")

    # ── Update Excel ──────────────────────────────────────────────
    print("\nUpdating Excel...")
    updated = update_excel(wb, ws, targets, EXCEL_PATH)
    print(f"  {updated} cell(s) filled (highlighted in green).")
    print(f"\nDone. File saved:\n  {EXCEL_PATH}")


if __name__ == '__main__':
    main()
