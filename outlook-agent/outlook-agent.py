"""
Outlook External Contacts Extractor
====================================
Connects to the local Outlook desktop app via COM automation and produces an
Excel file of external professional contacts — suppliers, partners, sellers,
distributors, and anyone outside SafeFields with whom communication was
two-way (i.e. they replied or initiated contact).

Contacts that were only ever recipients of outgoing mail (no reply back) are
excluded per the configured filter.

Requirements: pip install pywin32 openpyxl
Run: python outlook_contacts.py
"""

import os
import re
import sys
from collections import defaultdict
from datetime import datetime

# ── Dependency check ──────────────────────────────────────────────────────────
try:
    import win32com.client
    import pywintypes
except ImportError:
    sys.exit("ERROR: pywin32 not installed. Run: pip install pywin32")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  CONFIGURATION  — edit here if needed
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SAFEFIELDS_DOMAINS = [
    'safefields.com', 'safefields.co.uk', 'safefields.org',
    'safefields-tech.com',   # internal tech domain
]

# Full names to exclude (case-insensitive exact match)
EXCLUDED_NAMES = {
    'maxim birger',    # self
    'gal grossfeld',
}

# Email address substrings to exclude (case-insensitive)
EXCLUDED_EMAIL_KEYWORDS = [
    'donotreply', 'do-not-reply', 'do_not_reply',
    'noreply', 'no-reply', 'no_reply',
    'notifications@', 'notification@',
    'newsletter@', 'mailer@', 'support@',
]

# Domain fragments for known services/newsletters to exclude
EXCLUDED_SERVICE_DOMAINS = [
    'disneyplus.com', 'disney-plus.com', 'disney.com',
    'anydesk.com',
    'jlcpcb.com',
    'easyeda.com',
    'altium.com',
    'circuitlab.com',
    'leumi.co.il', 'bankleumi.co.il', 'leumi.com',
]

# Display name fragments for services (case-insensitive substring match)
EXCLUDED_DISPLAY_KEYWORDS = [
    'disney+', 'disney plus',
    'anydesk',
    'jlcpcb',
    'easyeda',
    'altium',
    'circuitlab academy',
    'bank leumi', 'leumi',
    'do not reply', 'donotreply', 'no-reply', 'no reply',
]

OUTPUT_DIR  = r'C:\Users\maxbi\OneDrive\Documents\GitHub\My-Agents\outlook-agent'
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'external_contacts.xlsx')

# Outlook folders to ignore while scanning received mail
SKIP_FOLDER_NAMES = {
    'deleted items', 'drafts', 'junk email', 'spam',
    'outbox', 'trash', 'clutter', 'conversation history',
}

# Outlook item/folder class constants
OL_MAIL_ITEM    = 43
OL_CONTACT_ITEM = 40
OL_FOLDER_INBOX    = 6
OL_FOLDER_SENT     = 5
OL_FOLDER_CONTACTS = 10

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  HELPERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
_EMAIL_RE = re.compile(r'[\w.+\-]+@[\w\-]+\.[\w.\-]+')


def extract_smtp(raw: str) -> str:
    """Pull a clean SMTP address from 'Name <addr>' or bare strings."""
    if not raw:
        return ''
    raw = raw.strip()
    m = re.search(r'<([^>]+)>', raw)
    if m:
        candidate = m.group(1).strip().lower()
        if '@' in candidate:
            return candidate
    m = _EMAIL_RE.search(raw)
    if m:
        return m.group().lower()
    return ''


def get_domain(email: str) -> str:
    return email.split('@')[-1].lower() if '@' in email else ''


def is_internal(email: str) -> bool:
    domain = get_domain(email)
    return any(
        domain == sf or domain.endswith('.' + sf)
        for sf in SAFEFIELDS_DOMAINS
    )


def is_excluded(email: str, display_name: str = '') -> bool:
    """Return True if this sender should be filtered out."""
    email_lc   = email.lower()
    name_lc    = display_name.strip().lower()
    domain     = get_domain(email_lc)

    # Excluded full names
    if name_lc and name_lc in EXCLUDED_NAMES:
        return True

    # No-reply / do-not-reply patterns in the email address
    if any(kw in email_lc for kw in EXCLUDED_EMAIL_KEYWORDS):
        return True

    # Known service domains
    if any(domain == sd or domain.endswith('.' + sd)
           for sd in EXCLUDED_SERVICE_DOMAINS):
        return True

    # Known service display-name keywords
    if name_lc and any(kw in name_lc for kw in EXCLUDED_DISPLAY_KEYWORDS):
        return True

    return False


def log(msg: str, end='\n'):
    print(msg, end=end, flush=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  OUTLOOK COM HELPERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def resolve_exchange_address(entry) -> str:
    """Resolve an Exchange/EX address entry to its SMTP address."""
    try:
        exch = entry.GetExchangeUser()
        if exch:
            return extract_smtp(exch.PrimarySmtpAddress)
    except Exception:
        pass
    try:
        return extract_smtp(entry.Address)
    except Exception:
        return ''


def get_sender_smtp(mail) -> str:
    """Return the SMTP address of the sender of a mail item."""
    try:
        if getattr(mail, 'SenderEmailType', '') == 'EX':
            try:
                sender_entry = mail.Sender
                if sender_entry:
                    addr = resolve_exchange_address(sender_entry)
                    if addr:
                        return addr
            except Exception:
                pass
        addr = getattr(mail, 'SenderEmailAddress', '')
        if addr and '@' in addr:
            return extract_smtp(addr)
    except Exception:
        pass
    return ''


def get_recipients_smtp(mail) -> list:
    """Return a list of SMTP recipient addresses for a mail item."""
    results = []
    try:
        recips = mail.Recipients
        for i in range(1, recips.Count + 1):
            try:
                recip = recips.Item(i)
                addr = getattr(recip, 'Address', '')
                if addr and '@' in addr and not addr.startswith('/O='):
                    results.append(extract_smtp(addr))
                else:
                    entry = getattr(recip, 'AddressEntry', None)
                    if entry:
                        resolved = resolve_exchange_address(entry)
                        if resolved:
                            results.append(resolved)
            except Exception:
                pass
    except Exception:
        pass
    return [r for r in results if r]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  FOLDER ITERATION
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def iter_mail_in_folder(folder, recurse=True, _depth=0, max_depth=6):
    """Yield every mail item in a folder (and optionally sub-folders)."""
    if _depth > max_depth:
        return
    try:
        items = folder.Items
        total = items.Count
        if total:
            log(f"    [{folder.Name}] {total} items", end=' ... ')
            yielded = 0
            for i in range(1, total + 1):
                try:
                    item = items.Item(i)
                    if getattr(item, 'Class', 0) == OL_MAIL_ITEM:
                        yield item
                        yielded += 1
                except Exception:
                    pass
            log(f"{yielded} mail items")
    except Exception as e:
        log(f"    [warn] Could not read '{getattr(folder, 'Name', '?')}': {e}")

    if recurse:
        try:
            for j in range(1, folder.Folders.Count + 1):
                try:
                    sub = folder.Folders.Item(j)
                    name_lower = sub.Name.lower()
                    if name_lower not in SKIP_FOLDER_NAMES:
                        yield from iter_mail_in_folder(
                            sub, recurse=True,
                            _depth=_depth + 1, max_depth=max_depth
                        )
                except Exception:
                    pass
        except Exception:
            pass


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  FULL-PROFILE FOLDER WALKER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Folders whose NAME marks them (and their children) as "sent mail"
SENT_FOLDER_NAMES = {'sent items', 'sent mail', 'sent', 'gesendete elemente'}

# Folders to skip entirely — noise / system / deletions
SKIP_FOLDER_NAMES = {
    'deleted items', 'recoverable items', 'drafts', 'outbox',
    'junk email', 'spam', 'trash', 'clutter',
    'conversation history', 'rss feeds', 'rss subscriptions',
    'quick step settings', 'sync issues', 'conflicts',
    'local failures', 'server failures',
}


def _absorb_contact(item, contacts_db: dict, folder_label: str):
    """Extract a single Outlook contact item into contacts_db."""
    try:
        email = ''
        for field in ('Email1Address', 'Email2Address', 'Email3Address'):
            raw = getattr(item, field, '') or ''
            if raw and '@' in raw:
                email = extract_smtp(raw)
                if email:
                    break
        if not email or is_internal(email):
            return
        name     = (getattr(item, 'FullName',              '') or '').strip()
        company  = (getattr(item, 'CompanyName',           '') or '').strip()
        position = (getattr(item, 'JobTitle',              '') or '').strip()
        phone    = (
            getattr(item, 'BusinessTelephoneNumber', '') or
            getattr(item, 'MobileTelephoneNumber',   '') or
            getattr(item, 'HomeTelephoneNumber',     '') or ''
        ).strip()
        if is_excluded(email, name):
            return
        # Merge — prefer richer records (more fields filled)
        existing = contacts_db.get(email)
        new_score = bool(name) + bool(company) + bool(position) + bool(phone)
        if existing is None or new_score > existing.get('_score', 0):
            contacts_db[email] = {
                'name': name, 'company': company,
                'position': position, 'phone': phone,
                'email': email, '_score': new_score,
                '_source': folder_label,
            }
    except Exception:
        pass


def _walk_folder(folder, received_from, sent_to, contacts_db,
                 is_sent_tree=False, depth=0, max_depth=12):
    """Recursively walk one folder, updating the three accumulators."""
    if depth > max_depth:
        return

    try:
        name_lower = folder.Name.strip().lower()
    except Exception:
        return

    if name_lower in SKIP_FOLDER_NAMES:
        return

    is_sent = is_sent_tree or (name_lower in SENT_FOLDER_NAMES)
    label   = folder.Name

    # ── Process items in this folder ─────────────────────────────────────────
    try:
        items = folder.Items
        count = items.Count
        if count:
            log(f"    [{label}] {count} item(s)", end=' ... ')
            mail_n = contact_n = 0
            for i in range(1, count + 1):
                try:
                    item       = items.Item(i)
                    item_class = getattr(item, 'Class', 0)

                    if item_class == OL_MAIL_ITEM:
                        if is_sent:
                            for addr in get_recipients_smtp(item):
                                if not is_internal(addr):
                                    sent_to.add(addr)
                        else:
                            smtp    = get_sender_smtp(item)
                            display = (getattr(item, 'SenderName', '') or '').strip()
                            if smtp and not is_internal(smtp) and not is_excluded(smtp, display):
                                existing = received_from.get(smtp, '')
                                if len(display) > len(existing):
                                    received_from[smtp] = display
                        mail_n += 1

                    elif item_class == OL_CONTACT_ITEM:
                        _absorb_contact(item, contacts_db, label)
                        contact_n += 1

                except Exception:
                    pass

            parts = []
            if mail_n:    parts.append(f"{mail_n} mail")
            if contact_n: parts.append(f"{contact_n} contacts")
            log(', '.join(parts) if parts else 'skipped')
    except Exception as e:
        log(f"    [warn] '{label}': {e}")

    # ── Recurse into sub-folders ──────────────────────────────────────────────
    try:
        for j in range(1, folder.Folders.Count + 1):
            try:
                _walk_folder(folder.Folders.Item(j),
                             received_from, sent_to, contacts_db,
                             is_sent_tree=is_sent,
                             depth=depth + 1, max_depth=max_depth)
            except Exception:
                pass
    except Exception:
        pass


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  MAIN SCAN
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def scan_outlook():
    """Walk every store / every folder in the MAPI profile."""
    log("Connecting to Outlook (must be running with a profile open)...")
    try:
        outlook = win32com.client.Dispatch('Outlook.Application')
        ns      = outlook.GetNamespace('MAPI')
    except Exception as e:
        sys.exit(f"ERROR: Cannot connect to Outlook - {e}\n"
                 "Make sure Outlook is open and a mail profile is loaded.")

    received_from: dict = {}
    sent_to:       set  = set()
    contacts_db:   dict = {}

    store_count = ns.Stores.Count
    log(f"\nFound {store_count} mail store(s) in profile.")

    for s in range(1, store_count + 1):
        try:
            store = ns.Stores.Item(s)
            store_name = getattr(store, 'DisplayName', f'Store {s}')
            log(f"\n-- Store: {store_name} --")
            root = store.GetRootFolder()
            _walk_folder(root, received_from, sent_to, contacts_db)
        except Exception as e:
            log(f"  [warn] Could not open store {s}: {e}")

    log(f"\nScan complete.")
    log(f"  External senders (received): {len(received_from)}")
    log(f"  External recipients (sent):  {len(sent_to)}")
    log(f"  Contacts with rich data:     {len(contacts_db)}")

    return received_from, sent_to, contacts_db


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  BUILD CONTACT LIST
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def build_contact_list(received_from, sent_to, contacts_db):
    """
    Inclusion rules:
      A. Anyone who emailed us (appears in received_from) — two-way confirmed.
      B. Anyone saved in a Contacts/Suppliers/people folder (contacts_db) —
         explicitly stored by the user, treat as intentional.
    One-way outgoing only (sent_to but not in received_from or contacts_db)
    are excluded.
    """
    rows = []
    seen = set()

    # Pool A: received mail senders
    for email, display_name in received_from.items():
        if email in seen:
            continue
        seen.add(email)

        if email in contacts_db:
            c = contacts_db[email].copy()
            # Enrich name from email display if contacts entry lacks one
            if not c.get('name'):
                c['name'] = display_name or email.split('@')[0].replace('.', ' ').title()
        else:
            name = display_name or email.split('@')[0].replace('.', ' ').title()
            c = {'email': email, 'name': name,
                 'company': '', 'position': '', 'phone': ''}

        c['email'] = email
        rows.append(c)

    # Pool B: saved contacts not already captured via email
    for email, c in contacts_db.items():
        if email in seen:
            continue
        seen.add(email)
        entry = c.copy()
        entry['email'] = email
        rows.append(entry)

    # Sort: company first, then name
    rows.sort(key=lambda x: (
        x.get('company', '').lower(),
        x.get('name', '').lower(),
    ))
    return rows


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  EXCEL EXPORT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
_BLUE_DARK  = '1F3864'
_BLUE_LIGHT = 'D6E4F0'
_WHITE      = 'FFFFFF'
_STRIPE     = 'EEF4FB'


def _thin_border():
    thin = Side(style='thin', color='C0C0C0')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def write_excel(contacts: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'External Contacts'

    # ── Header row ────────────────────────────────────────────────────────────
    columns = [
        ('Full Name',  30),
        ('Company',    32),
        ('Position',   28),
        ('Email',      36),
        ('Phone',      18),
    ]
    hdr_font  = Font(name='Calibri', bold=True, color=_WHITE, size=11)
    hdr_fill  = PatternFill('solid', fgColor=_BLUE_DARK)
    hdr_align = Alignment(horizontal='center', vertical='center')

    for col_idx, (hdr, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(columns))}1'

    # ── Data rows ─────────────────────────────────────────────────────────────
    alt_fill   = PatternFill('solid', fgColor=_STRIPE)
    data_font  = Font(name='Calibri', size=10)
    data_align = Alignment(vertical='center', wrap_text=False)
    email_font = Font(name='Calibri', size=10, color='0563C1', underline='single')

    for row_idx, c in enumerate(contacts, 2):
        values = [
            c.get('name',     ''),
            c.get('company',  ''),
            c.get('position', ''),
            c.get('email',    ''),
            c.get('phone',    ''),
        ]
        is_alt = (row_idx % 2 == 0)
        fill   = alt_fill if is_alt else None

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = data_align
            cell.border    = _thin_border()
            if col_idx == 4:           # email column — hyperlink style
                cell.font = email_font
            else:
                cell.font = data_font
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 18

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Run Info')
    info_rows = [
        ('Generated',       datetime.now().strftime('%Y-%m-%d  %H:%M')),
        ('Total contacts',  len(contacts)),
        ('Filter',          'External only · received at least one reply'),
        ('Excluded domains', ', '.join(SAFEFIELDS_DOMAINS)),
    ]
    bold = Font(bold=True, name='Calibri')
    norm = Font(name='Calibri')
    for r, (k, v) in enumerate(info_rows, 1):
        ws2.cell(r, 1, k).font  = bold
        ws2.cell(r, 2, str(v)).font = norm
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 45

    wb.save(output_path)
    log(f"\n  Saved -> {output_path}")
    log(f"  Total exported: {len(contacts)} contact(s)")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  ENTRY POINT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def main():
    log("=" * 60)
    log("  Outlook External Contacts Extractor")
    log("  SafeFields excluded | two-way communication only")
    log("=" * 60)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    received_from, sent_to, contacts_db = scan_outlook()

    log("\nBuilding contact list...")
    contacts = build_contact_list(received_from, sent_to, contacts_db)

    if not contacts:
        log("\nNo external contacts matched the criteria.")
        log("Check that Outlook is open and SAFEFIELDS_DOMAINS is correct.")
        return

    log(f"  {len(contacts)} contact(s) qualify.")

    log("\nWriting Excel file...")
    write_excel(contacts, OUTPUT_FILE)

    log("\nDone. Open the file to review:")
    log(f"  {OUTPUT_FILE}")


if __name__ == '__main__':
    main()
