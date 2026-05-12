import requests
import schedule
import time
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font

CITIES = {
    "Hadera":   {"lat": 32.4342, "lon": 34.9196},
    "Netanya":  {"lat": 32.3337, "lon": 34.8601},
    "Tel Aviv": {"lat": 32.0853, "lon": 34.7818},
}

EXCEL_FILE = "weather_data.xlsx"
INTERVAL_MINUTES = 5
DURATION_HOURS = 24


def fetch_weather(lat: float, lon: float) -> dict | None:
    url = (
        "https://api.open-meteo.com/v1/forecast"
        f"?latitude={lat}&longitude={lon}"
        "&current=temperature_2m,relative_humidity_2m,wind_speed_10m,wind_direction_10m"
    )
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        return response.json()["current"]
    except Exception as e:
        print(f"  Error fetching ({lat}, {lon}): {e}")
        return None


def build_headers() -> list[str]:
    headers = ["Timestamp"]
    for city in CITIES:
        headers += [
            f"{city} Temp (°C)",
            f"{city} Humidity (%)",
            f"{city} Wind Speed (km/h)",
            f"{city} Wind Dir (°)",
        ]
    return headers


def get_or_create_workbook():
    path = Path(EXCEL_FILE)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weather Data"
        ws.append(build_headers())
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.column_dimensions["A"].width = 20
    return wb, ws


def collect_and_save():
    now = datetime.now()
    print(f"\n[{now.strftime('%Y-%m-%d %H:%M:%S')}] Collecting weather data...")
    row = [now.strftime("%Y-%m-%d %H:%M:%S")]

    for city, coords in CITIES.items():
        data = fetch_weather(coords["lat"], coords["lon"])
        if data:
            temp     = data.get("temperature_2m")
            humidity = data.get("relative_humidity_2m")
            wspeed   = data.get("wind_speed_10m")
            wdir     = data.get("wind_direction_10m")
            print(f"  {city}: {temp}°C  |  {humidity}% humidity  |  {wspeed} km/h @ {wdir}°")
        else:
            temp = humidity = wspeed = wdir = None
            print(f"  {city}: N/A")
        row += [temp, humidity, wspeed, wdir]

    wb, ws = get_or_create_workbook()
    ws.append(row)
    wb.save(EXCEL_FILE)
    print(f"  Saved to {EXCEL_FILE}")


if __name__ == "__main__":
    end_time = datetime.now() + timedelta(hours=DURATION_HOURS)
    print("Weather monitoring agent started.")
    print(f"Cities: {', '.join(CITIES)}")
    print(f"Interval: every {INTERVAL_MINUTES} minutes")
    print(f"Will stop at: {end_time.strftime('%Y-%m-%d %H:%M:%S')} (after {DURATION_HOURS}h)\n")

    collect_and_save()

    schedule.every(INTERVAL_MINUTES).minutes.do(collect_and_save)

    while datetime.now() < end_time:
        schedule.run_pending()
        time.sleep(1)

    print("\n24-hour collection complete. Agent stopped.")
