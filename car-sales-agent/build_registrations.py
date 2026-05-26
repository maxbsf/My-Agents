import sys, os
os.environ['PYTHONIOENCODING'] = 'utf-8'
sys.stdout.reconfigure(encoding='utf-8')
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

df = pd.read_json('output/vehicle_2026_raw.json')

MONTH_MAP   = {202601:'January 2026', 202602:'February 2026', 202603:'March 2026', 202604:'April 2026'}
MONTH_SHORT = {202601:'Jan', 202602:'Feb', 202603:'Mar', 202604:'Apr'}
TYPE_MAP    = {'P':'Passenger Car', '3':'Motorcycle', 'M':'Commercial/Van'}

df['month_name']    = df['sgira_month'].map(MONTH_MAP)
df['type_name']     = df['sug_degem'].map(TYPE_MAP)
df['kinuy_mishari'] = df['kinuy_mishari'].fillna('')

# Styles
DARK_BLUE  = PatternFill('solid', fgColor='1F4E79')
MID_BLUE   = PatternFill('solid', fgColor='2E75B6')
LIGHT_BLUE = PatternFill('solid', fgColor='D6E4F0')
ORANGE     = PatternFill('solid', fgColor='ED7D31')
GRAY       = PatternFill('solid', fgColor='F2F2F2')
WHITE_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BOLD       = Font(name='Arial', bold=True, size=10)
NORMAL     = Font(name='Arial', size=10)
NOTE_FONT  = Font(name='Arial', italic=True, color='595959', size=8)
CENTER     = Alignment(horizontal='center', vertical='center')
LEFT       = Alignment(horizontal='left',   vertical='center')
NUM_FMT    = '#,##0'
SOURCE_NOTE = ('Source: Israeli Ministry of Transport via data.gov.il | '
               'Dataset: khamut kli rechev chadashim be-kod degem | '
               'Period: January-April 2026 | Published: May 3, 2026')

months_ordered = [202601, 202602, 202603, 202604]

wb = Workbook()

# ── Sheet 1: Raw Data ───────────────────────────────────────
ws1 = wb.active
ws1.title = 'Raw Data'

col_defs = [
    ('Month',           'sgira_month',    12),
    ('Month Name',      'month_name',     16),
    ('Vehicle Type',    'type_name',      16),
    ('Mfr Code',        'tozeret_cd',     10),
    ('Manufacturer',    'tozeret_nm',     24),
    ('Model Code',      'degem_cd',       10),
    ('Model ID',        'degem_nm',       16),
    ('Commercial Name', 'kinuy_mishari',  28),
    ('Units',           'car_num',        10),
]

for ci, (header, _, width) in enumerate(col_defs, 1):
    cell = ws1.cell(1, ci, header)
    cell.font = WHITE_FONT
    cell.fill = DARK_BLUE
    cell.alignment = CENTER
    ws1.column_dimensions[get_column_letter(ci)].width = width

for ri, row in enumerate(df.itertuples(), 2):
    vals = [row.sgira_month, row.month_name, row.type_name,
            row.tozeret_cd, row.tozeret_nm, row.degem_cd,
            row.degem_nm, row.kinuy_mishari, row.car_num]
    for ci, val in enumerate(vals, 1):
        cell = ws1.cell(ri, ci, val)
        cell.font = NORMAL
        cell.fill = GRAY if ri % 2 == 0 else PatternFill()
        cell.alignment = CENTER if ci in (1, 3, 4, 6, 9) else LEFT
        if ci == 9:
            cell.number_format = NUM_FMT

ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = 'A1:I' + str(len(df) + 1)
ws1.cell(len(df) + 3, 1, SOURCE_NOTE).font = NOTE_FONT
print('Sheet 1 (Raw Data): done')

# ── Sheet 2: Passenger Cars by Model x Month ───────────────
ws2 = wb.create_sheet('Passenger Cars by Month')
p = df[df['sug_degem'] == 'P'].copy()

pivot = (p.groupby(['tozeret_nm', 'kinuy_mishari', 'sgira_month'])['car_num']
          .sum().unstack(fill_value=0))
for m in months_ordered:
    if m not in pivot.columns:
        pivot[m] = 0
pivot = pivot[months_ordered]
pivot['Total'] = pivot.sum(axis=1)
pivot = pivot.sort_values('Total', ascending=False).reset_index()

headers2 = ['Manufacturer', 'Model'] + [MONTH_SHORT[m] for m in months_ordered] + ['Total Jan-Apr']
widths2  = [24, 28, 10, 10, 10, 10, 14]

for ci, (h, w) in enumerate(zip(headers2, widths2), 1):
    cell = ws2.cell(1, ci, h)
    cell.font = WHITE_FONT
    cell.fill = MID_BLUE if ci in (3, 4, 5, 6) else DARK_BLUE
    cell.alignment = CENTER
    ws2.column_dimensions[get_column_letter(ci)].width = w

for ri, row in enumerate(pivot.itertuples(index=False), 2):
    vals = [row.tozeret_nm, row.kinuy_mishari,
            row._2, row._3, row._4, row._5, row.Total]
    fill = GRAY if ri % 2 == 0 else PatternFill()
    for ci, val in enumerate(vals, 1):
        cell = ws2.cell(ri, ci, val)
        cell.font = NORMAL
        cell.fill = fill
        cell.alignment = LEFT if ci <= 2 else CENTER
        if ci >= 3:
            cell.number_format = NUM_FMT

tr2 = len(pivot) + 2
ws2.cell(tr2, 1, 'TOTAL').font = BOLD
ws2.cell(tr2, 1).fill = LIGHT_BLUE
for ci in range(3, 8):
    col_l = get_column_letter(ci)
    c = ws2.cell(tr2, ci, '=SUM(' + col_l + '2:' + col_l + str(tr2 - 1) + ')')
    c.font = BOLD
    c.fill = LIGHT_BLUE
    c.alignment = CENTER
    c.number_format = NUM_FMT

ws2.freeze_panes = 'C2'
ws2.auto_filter.ref = 'A1:G' + str(len(pivot) + 1)
ws2.cell(len(pivot) + 4, 1, SOURCE_NOTE).font = NOTE_FONT
print('Sheet 2 (Passenger Cars by Month): done')

# ── Sheet 3: Summary by Manufacturer ───────────────────────
ws3 = wb.create_sheet('By Manufacturer')

mfr = (p.groupby(['tozeret_nm', 'sgira_month'])['car_num']
        .sum().unstack(fill_value=0))
for m in months_ordered:
    if m not in mfr.columns:
        mfr[m] = 0
mfr = mfr[months_ordered]
mfr['Total'] = mfr.sum(axis=1)
mfr = mfr.sort_values('Total', ascending=False).reset_index()

headers3 = ['Manufacturer'] + [MONTH_SHORT[m] for m in months_ordered] + ['Total', 'Market Share']
widths3  = [26, 10, 10, 10, 10, 12, 13]

for ci, (h, w) in enumerate(zip(headers3, widths3), 1):
    cell = ws3.cell(1, ci, h)
    cell.font = WHITE_FONT
    cell.fill = MID_BLUE if ci in (2, 3, 4, 5) else DARK_BLUE
    cell.alignment = CENTER
    ws3.column_dimensions[get_column_letter(ci)].width = w

total_row = len(mfr) + 2
for ri, row in enumerate(mfr.itertuples(index=False), 2):
    vals = [row.tozeret_nm, row._1, row._2, row._3, row._4, row.Total]
    fill = GRAY if ri % 2 == 0 else PatternFill()
    for ci, val in enumerate(vals, 1):
        cell = ws3.cell(ri, ci, val)
        cell.font = NORMAL
        cell.fill = fill
        cell.alignment = LEFT if ci == 1 else CENTER
        if ci >= 2:
            cell.number_format = NUM_FMT
    share = ws3.cell(ri, 7, '=F' + str(ri) + '/F$' + str(total_row))
    share.font = NORMAL
    share.fill = fill
    share.alignment = CENTER
    share.number_format = '0.0%'

for ci in range(1, 7):
    col_l = get_column_letter(ci)
    c = ws3.cell(total_row, ci)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = ORANGE
    c.alignment = CENTER if ci > 1 else LEFT
    if ci == 1:
        c.value = 'TOTAL'
    else:
        c.value = '=SUM(' + col_l + '2:' + col_l + str(total_row - 1) + ')'
        c.number_format = NUM_FMT

ws3.freeze_panes = 'B2'
ws3.auto_filter.ref = 'A1:G' + str(len(mfr) + 1)
ws3.cell(len(mfr) + 4, 1, SOURCE_NOTE).font = NOTE_FONT
print('Sheet 3 (By Manufacturer): done')

out_path = 'output/israel_vehicle_registrations_2026.xlsx'
wb.save(out_path)
print('Saved:', out_path)
