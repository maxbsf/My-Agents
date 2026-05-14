"""
Excel handler: reads the source file, builds output with identical formatting
plus sales columns (Overall TAM, 2026..2020).

Strategy for preserving x14 conditional formatting (icon sets):
  1. Copy source xlsx byte-for-byte to output path.
  2. Open the copy with openpyxl, add new columns, save back in-place.
  3. openpyxl re-writes sheet1.xml but silently drops the x14 extLst CF block.
  4. Post-save: use zipfile to re-inject the original x14 extLst into sheet1.xml.
"""
import copy
import io
import logging
import re
import shutil
import zipfile
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

from config import (
    SOURCE_FILE, SOURCE_SHEET, OUTPUT_FILE,
    YEARS, TYPE_DISPLAY_MAP, ELECTRIFIED_TYPES
)

logger = logging.getLogger(__name__)


BLUE_FONT  = Font(name="Arial", color="0000FF", size=10)
BLACK_FONT = Font(name="Arial", color="000000", size=10)
HEADER_FILL_FALLBACK = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CENTER  = Alignment(horizontal="center", vertical="center")
NUM_FMT = "#,##0"


# ---------------------------------------------------------------------------
# Public: read models from source
# ---------------------------------------------------------------------------

def read_models(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb[SOURCE_SHEET]
    header_row = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] == "#":
            header_row = list(row)
            break
    if header_row is None:
        raise ValueError("Header row '#' not found")

    idx = {h: i for i, h in enumerate(header_row)}
    models = []
    for row in ws.iter_rows(min_row=ws.min_row + 1, values_only=False):
        num_cell = row[idx["#"]]
        if not isinstance(num_cell.value, (int, float)):
            continue
        raw_type = str(row[idx["Type"]].value or "").strip().upper()
        if raw_type not in ELECTRIFIED_TYPES:
            continue
        models.append({
            "row_num": int(num_cell.value),
            "vendor":  str(row[idx["Vendor"]].value or "").strip(),
            "model":   row[idx["Model"]].value,
            "type":    TYPE_DISPLAY_MAP.get(raw_type, raw_type),
            "max_mf":  row[idx["Max MF"]].value,
            "aae":     row[idx["Average Annual Exposure (AAE)"]].value,
        })

    logger.info(f"Loaded {len(models)} electrified models from source")
    return models


# ---------------------------------------------------------------------------
# Public: build output Excel
# ---------------------------------------------------------------------------

def build_output(sales_results: list[dict]) -> str:
    out_path = Path(OUTPUT_FILE)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Step 1 — extract x14 extLst block from source before openpyxl touches it
    x14_block = _extract_x14_extlst(SOURCE_FILE)

    # Step 2 — copy source to output (preserves all raw XML incl. x14 block)
    shutil.copy2(SOURCE_FILE, out_path)

    # Step 3 — open copy, add new columns, save in-place
    wb = load_workbook(str(out_path))
    ws = wb[SOURCE_SHEET]

    _normalise_mhev(ws)

    header_row     = _find_header_row(ws)
    data_start_row = header_row + 1

    # Detect existing TAM/year columns or create them
    tam_col, year_col_map = _resolve_sales_columns(ws, header_row)

    _write_sales_data_v2(ws, data_start_row, tam_col, year_col_map, sales_results)

    wb.save(str(out_path))
    logger.info(f"Saved: {out_path}")

    # Step 4 — re-inject the x14 extLst that openpyxl stripped
    if x14_block:
        _reinject_x14_extlst(str(out_path), x14_block)
        logger.info("x14 conditional formatting re-injected")
    else:
        logger.warning("No x14 extLst found in source — skipping CF re-injection")

    return str(out_path)


# ---------------------------------------------------------------------------
# Helpers: sheet manipulation
# ---------------------------------------------------------------------------

def _find_header_row(ws) -> int:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "#":
                return cell.row
    return 2


def _normalise_mhev(ws):
    type_col = None
    for row in ws.iter_rows(max_row=5):
        for cell in row:
            if cell.value == "Type":
                type_col = cell.column
                break
        if type_col:
            break
    if not type_col:
        return
    for row in ws.iter_rows(min_row=_find_header_row(ws) + 1):
        cell = row[type_col - 1]
        if str(cell.value or "").strip().upper() == "MHEV":
            cell.value = "HEV"



def _resolve_sales_columns(ws, header_row: int) -> tuple[int, dict]:
    """
    Find existing TAM and year columns by header text.
    Creates missing year columns (and TAM if absent) appended after max_column.
    Returns (tam_col, {year_int: col_int}).
    """
    tam_col = None
    year_cols = {}
    for cell in ws[header_row]:
        v = str(cell.value or "").strip()
        if not v:
            continue
        if "TAM" in v.upper():
            tam_col = cell.column
        else:
            try:
                yr = int(v.replace("(YTD)", "").replace("YTD", "").strip())
                if 2015 <= yr <= 2035:
                    year_cols[yr] = cell.column
            except ValueError:
                pass

    fill = _get_header_fill(ws, header_row)
    next_col = ws.max_column + 1

    if tam_col is None:
        cell = ws.cell(row=header_row, column=next_col, value="Overall TAM")
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(next_col)].width = 14
        tam_col = next_col
        next_col += 1

    for year in sorted(YEARS, reverse=True):
        if year not in year_cols:
            label = f"{year} (YTD)" if year == max(YEARS) else str(year)
            cell = ws.cell(row=header_row, column=next_col, value=label)
            cell.font = HEADER_FONT
            cell.fill = fill
            cell.alignment = CENTER
            ws.column_dimensions[get_column_letter(next_col)].width = 14
            year_cols[year] = next_col
            next_col += 1

    return tam_col, year_cols


def _write_sales_data_v2(ws, data_start_row: int, tam_col: int,
                         year_col_map: dict, results: list[dict]):
    by_rownum = {r["row_num"]: r["sales"] for r in results}
    today_str = date.today().isoformat()

    for row in ws.iter_rows(min_row=data_start_row):
        num_cell = row[0]
        if not isinstance(num_cell.value, (int, float)):
            continue
        row_num   = int(num_cell.value)
        excel_row = num_cell.row
        if row_num not in by_rownum:
            continue

        sales = by_rownum[row_num]

        for year, col in year_col_map.items():
            cell = ws.cell(row=excel_row, column=col)
            val  = sales.get(year)
            cell.value         = val
            cell.font          = BLUE_FONT
            cell.alignment     = CENTER
            cell.number_format = NUM_FMT
            if val is not None:
                cell.comment = Comment(
                    f"Source: carzone.co.il, scraped {today_str}",
                    "CarSalesAgent"
                )

        # TAM = sum of all year columns in this row
        year_cols_sorted = sorted(year_col_map.values())
        refs = "+".join(
            f"IFERROR({get_column_letter(c)}{excel_row},0)"
            for c in year_cols_sorted
        )
        tam_cell = ws.cell(row=excel_row, column=tam_col)
        tam_cell.value         = f"={refs}"
        tam_cell.font          = BLACK_FONT
        tam_cell.alignment     = CENTER
        tam_cell.number_format = NUM_FMT


def _get_header_fill(ws, header_row: int) -> PatternFill:
    for cell in ws[header_row]:
        if cell.fill and cell.fill.fgColor:
            rgb = cell.fill.fgColor.rgb
            if rgb not in ("00000000", "FFFFFFFF", None):
                return copy.copy(cell.fill)
    return HEADER_FILL_FALLBACK


# ---------------------------------------------------------------------------
# Helpers: x14 CF preservation via zipfile surgery
# ---------------------------------------------------------------------------

_X14_EXTLST_RE = re.compile(
    r'<extLst>.*?</extLst>',
    re.DOTALL
)
_X14_CF_RE = re.compile(
    r'<x14:conditionalFormattings>.*?</x14:conditionalFormattings>',
    re.DOTALL
)


def _extract_x14_extlst(xlsx_path: str) -> str | None:
    """Return the raw extLst XML string from sheet1.xml of the source file."""
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            xml = z.read('xl/worksheets/sheet1.xml').decode('utf-8')
        # Find extLst containing x14 CF
        for m in _X14_EXTLST_RE.finditer(xml):
            block = m.group(0)
            if 'x14:conditionalFormatting' in block:
                return block
    except Exception as e:
        logger.warning(f"Could not extract x14 extLst: {e}")
    return None


def _reinject_x14_extlst(xlsx_path: str, x14_block: str):
    """
    Re-insert the x14 extLst block into sheet1.xml of the already-saved output file.
    openpyxl strips this on save; we add it back just before </worksheet>.
    """
    try:
        # Read all files from the saved xlsx into memory
        buf = io.BytesIO()
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            with zipfile.ZipFile(buf, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename == 'xl/worksheets/sheet1.xml':
                        xml = data.decode('utf-8')
                        # Remove any partial extLst openpyxl may have written (safety)
                        xml = _X14_EXTLST_RE.sub('', xml)
                        # Inject our original block right before </worksheet>
                        xml = xml.replace('</worksheet>', x14_block + '</worksheet>')
                        data = xml.encode('utf-8')
                    zout.writestr(item, data)

        # Write back atomically
        with open(xlsx_path, 'wb') as f:
            f.write(buf.getvalue())

    except Exception as e:
        logger.warning(f"CF re-injection failed: {e}")
