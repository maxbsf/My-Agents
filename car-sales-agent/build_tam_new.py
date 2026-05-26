import sys, os
os.environ['PYTHONIOENCODING'] = 'utf-8'
sys.stdout.reconfigure(encoding='utf-8')

import io
import json
import math
import re
import zipfile
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Load gov registration data ────────────────────────────
with open('output/vehicle_2020_2025_raw.json', encoding='utf-8') as f:
    recs_2020_2025 = json.load(f)
with open('output/vehicle_2026_raw.json', encoding='utf-8') as f:
    recs_2026 = json.load(f)

all_recs = recs_2020_2025 + recs_2026
df = pd.DataFrame(all_recs)
df = df[df['sug_degem'] == 'P'].copy()
df['sgira_month'] = df['sgira_month'].astype(int)
df['car_num']     = pd.to_numeric(df['car_num'], errors='coerce').fillna(0).astype(int)
df['year']        = df['sgira_month'] // 100
df['kinuy']       = df['kinuy_mishari'].fillna('').str.strip().str.upper()
print(f'Total passenger car records 2020-2026: {len(df)}')

pivot = df.groupby(['kinuy', 'year', 'sgira_month'])['car_num'].sum()

def get_year_total(keys, year):
    total = 0
    for k in keys:
        k_up = k.strip().upper()
        try:
            total += pivot[k_up][year].sum()
        except KeyError:
            pass
    return total if total > 0 else None

def get_month_total(keys, month_int):
    total = 0
    for k in keys:
        k_up = k.strip().upper()
        try:
            total += int(pivot[k_up].get(month_int // 100, {}).get(month_int, 0))
        except (KeyError, TypeError):
            pass
    return total if total > 0 else None

# ── Manual overrides for rows with no source data ────────
MANUAL_TAM = {22: 956, 58: 1703, 59: 0}

# ── Load carzone data from cars_tam.xlsx ─────────────────
df_old = pd.read_excel('output/cars_tam.xlsx', sheet_name='Simplified', header=1)
CZ_YR_COLS = ['2026 (YTD)', '2025', '2024', '2023', '2022', '2021', '2020']
CZ_YR_MAP  = {'2026 (YTD)': 2026, '2025': 2025, '2024': 2024,
               '2023': 2023,      '2022': 2022, '2021': 2021, '2020': 2020}

def nanval(v):
    return v is None or (isinstance(v, float) and math.isnan(v))

def clean(v):
    return '' if nanval(v) else v

def safe_int(v):
    if nanval(v):
        return None
    return int(v)

# carzone_yr_map[row_num] = {2026: v_or_None, 2025: v, ..., 2020: v}
carzone_yr_map = {}
for _, row in df_old.iterrows():
    rn = int(row['#'])
    carzone_yr_map[rn] = {
        CZ_YR_MAP[c]: (None if nanval(row[c]) else int(row[c]))
        for c in CZ_YR_COLS
    }

df_old['OldTAM'] = df_old[CZ_YR_COLS].sum(axis=1, min_count=1)
old_tam_total_map = {
    int(row['#']): (None if nanval(row['OldTAM']) else int(row['OldTAM']))
    for _, row in df_old.iterrows()
}

# Max MF and AAE maps
maxmf_map = {int(row['#']): row['Max MF']                        for _, row in df_old.iterrows()}
aae_map   = {int(row['#']): row['Average Annual Exposure (AAE)'] for _, row in df_old.iterrows()}

# ── Model list ────────────────────────────────────────────
df_tam = pd.read_excel('output/cars_tam.xlsx', sheet_name='Simplified', header=1)

# ── MAPPING ──────────────────────────────────────────────
MAPPING = {
    1:  ['TIGGO8PRO PHEV'],
    2:  ['WRANGLER', 'WRANGLER UNLIMI', 'WRANGLER UNLIM', 'JEEP WRANGLER', 'WRANGLER RUBICO'],
    3:  ['JAECOO8 PHEV'],
    4:  ['HS HYBRID'],
    5:  ['OMODA 9 PHEV'],
    6:  ['ELANTRA HEV'],
    7:  ['OMODA 7 PHEV'],
    8:  ['JAECOO7 PHEV'],
    9:  ['JAECOO 5 HEV'],
    10: ['EHS', 'EHS7', 'EHS5'],
    11: ['MODEL 3'],
    12: None,
    13: ['SEAL U', 'BYD SEAL U'],
    14: ['PANAMERA 4 E-HY', 'PANAMERA 4 E HY', 'PANAMERA 4E HYB', 'PANAMERA4SE HYB'],
    15: ['VITARA'],
    16: ['SWIFT'],
    17: ['C10'],
    18: None,
    19: None,
    20: None,
    21: ['CAMRY HYBRID', 'CAMRY', 'CAMRY HYBRIDE', 'CAMRY HEV'],
    22: None,
    23: ['OUTLANDER'],
    24: ['NIRO HEV'],
    25: None,
    26: ['COROLLA SDN HSD', 'COROLLA HEV', 'COROLLA HYBRID'],
    27: ['RAV4 HYBRID', 'RAV4 HSD', 'RAV4'],
    28: ['MODEL Y'],
    29: ['X5 XDRIVE50E', 'X5 XDRIVE 50E', 'X5 XDIVE50E'],
    30: ['MG4'],
    31: None,
    32: ['KONA HYBRID'],
    33: None,
    34: None,
    35: ['530E', '530E XDRIVE'],
    36: ['XC40 B4'],
    37: ['JUKE HYBRID'],
    38: ['LYNKCO01 PHEV'],
    39: ['EX5'],
    40: ['SONATA HYBRID'],
    41: None,
    42: ['YARIS HYBRID'],
    43: ['YARIS CROSS HSD', 'YARIS CROSS HEV', 'YARIS CROSS', 'YARIS CROSS HYB'],
    44: ['G6'],
    45: ['BYD SEALION 7'],
    46: ['ZEEKR X'],
    47: None,
    48: None,
    49: None,
    50: None,
    51: ['ENYAQ 85'],
    52: None,
    53: None,
    54: ['C10'],
    55: None,
    56: None,
    57: ['ZS'],
    58: None,
    59: None,
    60: ['CHR', 'CHR PHEV'],
    61: ['EX30 SM', 'EX30 TM'],
    62: ['ZEEKR 001'],
    63: ['ZEEKR 7X'],
}

NOTES = {
    2:  'All Wrangler variants combined',
    10: 'All EHS variants combined (EHS/EHS5/EHS7)',
    13: 'SEAL U + BYD SEAL U combined',
    14: 'All Panamera PHEV variants combined',
    17: 'C10 covers both PHEV and BEV (rows 17 & 54)',
    21: 'All Camry HEV variants combined',
    26: 'Corolla sedan HEV only (excl. Corolla Cross)',
    27: 'RAV4 Hybrid/HSD combined',
    28: 'MODEL Y (all variants)',
    29: 'X5 xDrive50e PHEV only',
    36: 'XC40 B4 — mild hybrid',
    42: 'Yaris Hybrid (excl. Yaris Cross)',
    43: 'All Yaris Cross HEV variants combined',
    54: 'C10 covers both PHEV and BEV (rows 17 & 54)',
    60: 'CHR + CHR PHEV combined',
    61: 'EX30 SM + TM combined',
}

YEARS_DESC  = [2026, 2025, 2024, 2023, 2022, 2021, 2020]
MONTHS_2026 = [202601, 202602, 202603, 202604]
MONTH_SHORT = {202601: 'Jan', 202602: 'Feb', 202603: 'Mar', 202604: 'Apr'}

# ── Build data rows with merge logic ─────────────────────
rows = []
for _, tam_row in df_tam.iterrows():
    row_num = int(tam_row['#'])
    mapping = MAPPING.get(row_num)

    # Gov yearly values
    if mapping:
        gov_yr = {yr: get_year_total(mapping, yr) for yr in YEARS_DESC}
    else:
        gov_yr = {yr: None for yr in YEARS_DESC}

    gov_tam = sum(v for v in gov_yr.values() if not nanval(v))

    # Carzone yearly values and total
    cz_yr  = carzone_yr_map.get(row_num, {yr: None for yr in YEARS_DESC})
    old_tam = old_tam_total_map.get(row_num)
    cz_tam  = old_tam or 0

    # ── Merge: keep the source with the larger total ──────
    if cz_tam > gov_tam:
        eff_yr  = cz_yr
        eff_tam = old_tam
        source  = 'carzone'
    elif gov_tam > 0:
        eff_yr  = gov_yr
        eff_tam = gov_tam
        source  = 'gov'
    else:
        eff_yr  = {yr: None for yr in YEARS_DESC}
        # Apply manual override if present
        manual  = MANUAL_TAM.get(row_num)
        eff_tam = manual  # may be 0 or a real number or None
        source  = 'none'

    # Monthly 2026 breakdown (always from gov when available, else None)
    if mapping:
        month_vals = {m: get_month_total(mapping, m) for m in MONTHS_2026}
    else:
        month_vals = {m: None for m in MONTHS_2026}

    row = {
        '#':       row_num,
        'Vendor':  tam_row['Vendor'],
        'Model':   tam_row['Model'],
        'Type':    tam_row['Type'],
        'MaxMF':   maxmf_map.get(row_num),
        'AAE':     aae_map.get(row_num),
        'Notes':   NOTES.get(row_num, ''),
        'source':  source,
        'eff_tam': eff_tam,
        'gov_tam': gov_tam,
        'old_tam': old_tam,
    }
    for yr in YEARS_DESC:
        row[str(yr)] = eff_yr.get(yr)
    for m in MONTHS_2026:
        row[MONTH_SHORT[m]] = month_vals[m]

    rows.append(row)

# ── Extra rows not in the original TAM file ──────────────
EXTRA = [
    {
        '#': 64, 'Vendor': 'Audi',    'Model': 'Q4 E-TRON', 'Type': 'BEV',
        'MaxMF': 13.0, 'AAE': None,
        'Notes': 'All Q4 E-TRON variants combined (incl. Sportback)',
        'source': 'gov', 'gov_tam': 1636, 'old_tam': None,
        'eff_tam': 1636,
        '2026': 5,    '2025': 413,  '2024': 664,  '2023': 417,
        '2022': 137,  '2021': None, '2020': None,
        'Jan': 3, 'Feb': 1, 'Mar': 1, 'Apr': None,
    },
    {
        '#': 65, 'Vendor': 'Ford',    'Model': 'Puma', 'Type': 'ICE',
        'MaxMF': 8.3, 'AAE': None,
        'Notes': '',
        'source': 'gov', 'gov_tam': 1435, 'old_tam': None,
        'eff_tam': 1435,
        '2026': None, '2025': None, '2024': 317,  '2023': 184,
        '2022': 228,  '2021': 706,  '2020': None,
        'Jan': None, 'Feb': None, 'Mar': None, 'Apr': None,
    },
    {
        '#': 66, 'Vendor': 'Hyundai', 'Model': 'SantaFe', 'Type': 'ICE',
        'MaxMF': 5.4, 'AAE': None,
        'Notes': 'ICE only (excl. SantaFe HEV/PHEV)',
        'source': 'gov', 'gov_tam': 1540, 'old_tam': None,
        'eff_tam': 1540,
        '2026': 2,    '2025': 7,    '2024': 43,   '2023': 230,
        '2022': 449,  '2021': 495,  '2020': 314,
        'Jan': None, 'Feb': 1, 'Mar': None, 'Apr': 1,
    },
]
rows.extend(EXTRA)

# ── Sort by MaxMF descending (AAE proxy); None → bottom ──
rows.sort(key=lambda r: r['MaxMF'] if r['MaxMF'] is not None else -1,
          reverse=True)

# ── Renumber sequentially after sort ─────────────────────
for i, r in enumerate(rows, 1):
    r['#'] = i

df_out = pd.DataFrame(rows)

# ── Styles ────────────────────────────────────────────────
DARK_BLUE    = PatternFill('solid', fgColor='1F4E79')
MID_BLUE     = PatternFill('solid', fgColor='2E75B6')
ORANGE       = PatternFill('solid', fgColor='ED7D31')
GRAY         = PatternFill('solid', fgColor='F2F2F2')
NO_DATA_FILL = PatternFill('solid', fgColor='FFF2CC')   # yellow: no data
WHITE_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BOLD         = Font(name='Arial', bold=True, size=10)
BOLD_WHITE   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
NORMAL       = Font(name='Arial', size=10)
GRAY_FONT    = Font(name='Arial', size=9, color='808080', italic=True)
NOTE_FONT    = Font(name='Arial', italic=True, color='595959', size=8)
CENTER       = Alignment(horizontal='center', vertical='center')
LEFT         = Alignment(horizontal='left',   vertical='center')
NUM_FMT      = '#,##0'

SOURCE_NOTE = ('Source: Israeli Ministry of Transport via data.gov.il | '
               'Period: January 2020–April 2026 | Published: May 3, 2026 | '
               'Carzone: carzone.co.il scrape (May 2026)')

def write_header_row(ws, headers, widths, ytd_cols=None):
    ytd_cols = ytd_cols or set()
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, ci, h)
        cell.font = WHITE_FONT
        cell.fill = MID_BLUE if ci in ytd_cols else DARK_BLUE
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(ci)].width = w

# ── Column layout ─────────────────────────────────────────
# 1:#  2:Vendor  3:Model  4:Type  5:TAM  6:2026  7:2025  8:2024  9:2023
# 10:2022  11:2021  12:2020  13:OldTAM  14:Diff%  15:Notes
COL_MAXMF  = 5
COL_AAE    = 6
COL_TAM    = 7
COL_2026   = 8
COL_2025   = 9
COL_2024   = 10
COL_2023   = 11
COL_2022   = 12
COL_2021   = 13
COL_2020   = 14
COL_NOTES  = 15

YEAR_COLS = {2026: COL_2026, 2025: COL_2025, 2024: COL_2024,
             2023: COL_2023, 2022: COL_2022, 2021: COL_2021, 2020: COL_2020}

headers1 = ['#', 'Vendor', 'Model', 'Type',
            'Max MF', 'AAE',
            'TAM',
            '2026 (YTD)', '2025', '2024', '2023', '2022', '2021', '2020',
            'Notes']
widths1  = [4, 13, 24, 6,
            8, 8,
            12,
            11, 11, 11, 11, 11, 11, 11,
            36]

# ── Sheet 1: TAM by Year ──────────────────────────────────
wb = Workbook()
ws1 = wb.active
ws1.title = 'TAM by Year'
ws1.row_dimensions[1].height = 28

write_header_row(ws1, headers1, widths1, ytd_cols={COL_2026})

for ri in range(2, len(df_out) + 2):
    r       = df_out.iloc[ri - 2]
    source  = r['source']
    eff_tam = r['eff_tam']
    gov_tam = r['gov_tam']
    old_tam = r['old_tam']

    no_data   = (source == 'none')
    base_fill = NO_DATA_FILL if no_data else (GRAY if ri % 2 == 0 else PatternFill())

    def cell(ci, val, fmt=None, align=CENTER, gray_ok=False):
        v = clean(val)
        c = ws1.cell(ri, ci, v)
        c.fill      = base_fill
        c.alignment = align
        c.font      = GRAY_FONT if (no_data and gray_ok) else NORMAL
        if fmt and v != '':
            c.number_format = fmt
        return c

    cell(1, r['#'])
    cell(2, r['Vendor'], align=LEFT)
    cell(3, r['Model'],  align=LEFT)
    cell(4, r['Type'])

    # Max MF — always show; AAE computed as formula =IF(E="","",0.11*E+0.9)
    cell(COL_MAXMF, r['MaxMF'], fmt='0.0')
    E = get_column_letter(COL_MAXMF)
    aae_cell               = ws1.cell(ri, COL_AAE, f'=IF({E}{ri}="","",0.11*{E}{ri}+0.9)')
    aae_cell.fill          = base_fill
    aae_cell.alignment     = CENTER
    aae_cell.font          = NORMAL
    aae_cell.number_format = '0.0'

    # TAM — show manual value for 'none' rows, gray years only
    cell(COL_TAM, eff_tam, fmt=NUM_FMT)

    # Year columns — gray out on no-data rows
    for yr, col in YEAR_COLS.items():
        cell(col, r[str(yr)], fmt=NUM_FMT, gray_ok=True)

    cell(COL_NOTES, r['Notes'], align=LEFT)

# ── Summary rows ─────────────────────────────────────────
data_end  = len(df_out) + 1   # last data row index
r_tam     = data_end + 1      # TAM (overall)
r_sam     = data_end + 2      # SAM (best fit, AAE > 4)
r_som     = data_end + 3      # SOM (8% × SAM)

SALES_COLS = [COL_TAM, COL_2026, COL_2025, COL_2024,
              COL_2023, COL_2022, COL_2021, COL_2020]

AAE_COL_L  = get_column_letter(COL_AAE)   # F
AAE_RANGE  = f'${AAE_COL_L}$2:${AAE_COL_L}${data_end}'

FILL_TAM   = PatternFill('solid', fgColor='FABB6E')   # light orange
FILL_SAM   = PatternFill('solid', fgColor='E07B2A')   # medium orange
FILL_SOM   = PatternFill('solid', fgColor='B84D0A')   # dark orange

def summary_label(ws, ri, label, fill):
    c = ws.cell(ri, 1, label)
    c.font = BOLD_WHITE; c.fill = fill; c.alignment = LEFT
    # Extend fill across non-numeric columns B–F
    for ci in range(2, COL_TAM):
        ec = ws.cell(ri, ci)
        ec.fill = fill

def summary_cell(ws, ri, ci, formula, fill, fmt=NUM_FMT):
    c = ws.cell(ri, ci, formula)
    c.font = BOLD_WHITE; c.fill = fill
    c.alignment = CENTER; c.number_format = fmt

# TAM (overall) — SUM of all data rows
summary_label(ws1, r_tam, 'TAM (overall)', FILL_TAM)
for ci in SALES_COLS:
    col_l = get_column_letter(ci)
    summary_cell(ws1, r_tam, ci, f'=SUM({col_l}2:{col_l}{data_end})', FILL_TAM)

# SAM (best fit) — SUMIF AAE > 4
summary_label(ws1, r_sam, 'SAM (best fit)', FILL_SAM)
for ci in SALES_COLS:
    col_l  = get_column_letter(ci)
    sum_range = f'{col_l}2:{col_l}{data_end}'
    summary_cell(ws1, r_sam, ci,
                 f'=SUMIF({AAE_RANGE},">"&4,{sum_range})', FILL_SAM)

# SOM (obtainable) — 8% × SAM
summary_label(ws1, r_som, 'SOM (obtainable)', FILL_SOM)
for ci in SALES_COLS:
    sam_col_l = get_column_letter(ci)
    summary_cell(ws1, r_som, ci, f'=0.08*{sam_col_l}{r_sam}', FILL_SOM)

ws1.freeze_panes = 'E2'   # freeze #, Vendor, Model, Type
ws1.auto_filter.ref = f'B1:{get_column_letter(COL_NOTES)}{data_end}'

# Legend
leg = r_som + 2
ws1.cell(leg,   1, 'Legend:').font = BOLD
ws1.cell(leg+1, 1).fill = NO_DATA_FILL
ws1.cell(leg+1, 2, 'Yellow = No data from either source (gov or Carzone)').font = NOTE_FONT
ws1.cell(leg+2, 2, 'TAM = max(Gov registrations, Carzone) — yearly columns follow the larger source').font = NOTE_FONT
ws1.cell(leg+3, 2, 'AAE = Average Annual Exposure (= 0.11 × Max MF + 0.9) — traffic-light icons: green ≥ 7, yellow 4–7, red < 4').font = NOTE_FONT
ws1.cell(leg+4, 2, 'SAM = models with AAE > 4.0 | SOM = 8% × SAM').font = NOTE_FONT
ws1.cell(leg+5, 2, SOURCE_NOTE).font = NOTE_FONT

print('Sheet 1 (TAM by Year): done')

# ── Sheet 2: 2026 Monthly ─────────────────────────────────
ws2 = wb.create_sheet('2026 Monthly')

month_short_list = [MONTH_SHORT[m] for m in MONTHS_2026]
headers2 = ['#', 'Vendor', 'Model', 'Type'] + month_short_list + ['Total Jan-Apr', 'Notes']
widths2  = [4, 13, 24, 6] + [10]*4 + [13, 40]

write_header_row(ws2, headers2, widths2, ytd_cols={5, 6, 7, 8})

for ri in range(2, len(df_out) + 2):
    r      = df_out.iloc[ri - 2]
    source = r['source']
    has_monthly = any(not nanval(r[MONTH_SHORT[m]]) for m in MONTHS_2026)

    no_data_row = (source == 'none' and not has_monthly)
    bfill = NO_DATA_FILL if no_data_row else (GRAY if ri % 2 == 0 else PatternFill())

    for ci, val in enumerate([r['#'], r['Vendor'], r['Model'], r['Type'],
                               r['Jan'], r['Feb'], r['Mar'], r['Apr']], 1):
        c = ws2.cell(ri, ci, clean(val))
        c.fill = bfill; c.font = NORMAL
        c.alignment = LEFT if ci in (2, 3) else CENTER
        if ci >= 5 and not nanval(val):
            c.number_format = NUM_FMT

    total_c = ws2.cell(ri, 9)
    total_c.fill = bfill; total_c.alignment = CENTER; total_c.font = NORMAL
    if has_monthly:
        total_c.value         = f'=SUM(E{ri}:H{ri})'
        total_c.number_format = NUM_FMT

    nc = ws2.cell(ri, 10, r['Notes'])
    nc.fill = bfill; nc.alignment = LEFT; nc.font = NORMAL

ws2.column_dimensions['J'].width = 40

tr2 = len(df_out) + 2
ws2.cell(tr2, 1, 'TOTAL').fill  = ORANGE
ws2.cell(tr2, 1).font           = BOLD_WHITE
ws2.cell(tr2, 1).alignment      = CENTER
for ci in range(5, 10):
    col_l = get_column_letter(ci)
    c = ws2.cell(tr2, ci, f'=SUM({col_l}2:{col_l}{tr2-1})')
    c.font = BOLD_WHITE; c.fill = ORANGE; c.alignment = CENTER; c.number_format = NUM_FMT

ws2.freeze_panes = 'E2'
ws2.auto_filter.ref = 'A1:J' + str(len(df_out) + 1)

leg2 = tr2 + 2
ws2.cell(leg2,   1, 'Legend:').font = BOLD
ws2.cell(leg2+1, 1).fill = NO_DATA_FILL
ws2.cell(leg2+1, 2, 'Yellow = No data from either source (gov or Carzone)').font = NOTE_FONT
ws2.cell(leg2+3, 2, SOURCE_NOTE).font = NOTE_FONT

print('Sheet 2 (2026 Monthly): done')

# ── Save ──────────────────────────────────────────────────
out_path = 'output/cars_tam_new.xlsx'
wb.save(out_path)
print(f'Saved: {out_path}')

# ── Re-inject x14 traffic-light icon CF from cars_tam.xlsx ─
def _get_x14_block(src_path, new_sqref):
    """Extract the x14 extLst block from src and update its sqref."""
    with zipfile.ZipFile(src_path, 'r') as z:
        xml = z.read('xl/worksheets/sheet1.xml').decode('utf-8')
    m = re.search(r'<extLst>.*?</extLst>', xml, re.DOTALL)
    if not m:
        return None
    block = m.group(0)
    if 'x14:conditionalFormatting' not in block:
        return None
    # Replace the sqref range with the new one
    block = re.sub(r'<xm:sqref>[^<]+</xm:sqref>',
                   f'<xm:sqref>{new_sqref}</xm:sqref>', block)
    return block

def _inject_x14(xlsx_path, x14_block):
    """Re-insert x14 extLst into sheet1.xml just before </worksheet>."""
    buf = io.BytesIO()
    extlst_re = re.compile(r'<extLst>.*?</extLst>', re.DOTALL)
    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        with zipfile.ZipFile(buf, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == 'xl/worksheets/sheet1.xml':
                    xml = data.decode('utf-8')
                    xml = extlst_re.sub('', xml)  # remove any existing extLst
                    xml = xml.replace('</worksheet>', x14_block + '</worksheet>')
                    data = xml.encode('utf-8')
                zout.writestr(item, data)
    with open(xlsx_path, 'wb') as f:
        f.write(buf.getvalue())

# AAE CF applies to data rows only (row 2 → data_end)
data_last_row = data_end
x14_block = _get_x14_block('output/cars_tam.xlsx', f'F2:F{data_last_row}')
if x14_block:
    _inject_x14(out_path, x14_block)
    print(f'Traffic-light icon CF re-injected (AAE column F2:F{data_last_row})')
else:
    print('Warning: x14 block not found in cars_tam.xlsx')
print(f'Rows: {len(df_out)}')

gov_cnt = sum(1 for r in rows if r['source'] == 'gov')
cz_cnt  = sum(1 for r in rows if r['source'] == 'carzone')
no_cnt  = sum(1 for r in rows if r['source'] == 'none')
print(f'Source: gov={gov_cnt}  carzone={cz_cnt}  none={no_cnt}')

print('\nRows where Carzone data was used (Old TAM > Gov):')
for r in rows:
    if r['source'] == 'carzone':
        g = r['gov_tam']; o = r['old_tam'] or 0
        print(f'  Row {r["#"]:2d} {str(r["Model"])[:28]:28s}  Gov={g:7,.0f}  Carzone={o:7,.0f}')
