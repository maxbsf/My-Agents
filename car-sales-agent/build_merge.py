import sys, os
os.environ['PYTHONIOENCODING'] = 'utf-8'
sys.stdout.reconfigure(encoding='utf-8')
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Load source files ──────────────────────────────────────
df_tam = pd.read_excel('output/cars_tam_2026_dates.xlsx', header=1)
df_reg = pd.read_excel('output/israel_vehicle_registrations_2026.xlsx',
                       sheet_name='Passenger Cars by Month')

# Build lookup: model_name_upper → {Jan, Feb, Mar, Apr, Total}
reg_lookup = {}
for _, row in df_reg.iterrows():
    key = str(row['Model']).strip().upper()
    reg_lookup[key] = {
        'Jan': row['Jan'] or 0,
        'Feb': row['Feb'] or 0,
        'Mar': row['Mar'] or 0,
        'Apr': row['Apr'] or 0,
        'GovTotal': row['Total Jan-Apr'] or 0,
    }

def get_months(model_keys):
    """Sum monthly values across a list of model name keys."""
    result = {'Jan': 0, 'Feb': 0, 'Mar': 0, 'Apr': 0, 'GovTotal': 0}
    for k in model_keys:
        k_up = k.strip().upper()
        if k_up in reg_lookup:
            for col in result:
                result[col] += reg_lookup[k_up][col]
    return result

NONE = {'Jan': None, 'Feb': None, 'Mar': None, 'Apr': None, 'GovTotal': None}

# ── Explicit mapping: TAM row# → list of registration model keys ──
# None means no data in Jan-Apr 2026
MAPPING = {
    1:  ['TIGGO8PRO PHEV'],
    2:  ['WRANGLER', 'WRANGLER UNLIMI', 'WRANGLER UNLIM', 'JEEP WRANGLER', 'WRANGLER RUBICO'],
    3:  ['JAECOO8 PHEV'],
    4:  ['HS HYBRID'],
    5:  ['OMODA 9 PHEV'],
    6:  ['ELANTRA HEV'],
    7:  ['OMODA 7 PHEV'],
    8:  ['JAECOO7 PHEV'],
    9:  ['JAECOO 5 HEV'],
    10: ['EHS', 'EHS7', 'EHS5'],
    11: ['MODEL 3'],
    12: None,   # Ioniq 4 HEV — discontinued, no 2026 sales
    13: ['SEAL U', 'BYD SEAL U'],
    14: ['PANAMERA 4 E-HY', 'PANAMERA 4 E HY', 'PANAMERA 4E HYB', 'PANAMERA4SE HYB'],
    15: ['VITARA'],
    16: ['SWIFT'],
    17: ['C10'],              # LeapMotor C10 PHEV (shared entry with row 54)
    18: None,   # Audi Q8 e-tron — not in 2026 data
    19: None,   # KIA Niro Gen2 PHEV — no 2026 registrations
    20: None,   # KIA Niro Plus Gen1 PHEV — no 2026 registrations
    21: ['CAMRY HYBRID', 'CAMRY', 'CAMRY HYBRIDE', 'CAMRY HEV'],
    22: None,   # Chevrolet Silverado — not in 2026
    23: ['OUTLANDER'],
    24: ['NIRO HEV'],
    25: None,   # VOYAH Courage — not in 2026 data
    26: ['COROLLA SDN HSD', 'COROLLA HEV', 'COROLLA HYBRID'],   # sedan only, not cross
    27: ['RAV4 HYBRID', 'RAV4 HSD', 'RAV4'],
    28: ['MODEL Y'],
    29: ['X5 XDRIVE50E', 'X5 XDRIVE 50E', 'X5 XDIVE50E'],     # PHEV variants only
    30: ['MG4'],
    31: None,   # Seres M5 — not in 2026 data
    32: ['KONA HYBRID'],
    33: None,   # Seres M5 awd+rwd — not in 2026 data
    34: None,   # Hyundai Kona BEV — not in 2026 data
    35: ['530E', '530E XDRIVE'],
    36: ['XC40 B4'],
    37: ['JUKE HYBRID'],
    38: ['LYNKCO01 PHEV'],
    39: ['EX5'],
    40: ['SONATA HYBRID'],
    41: None,   # MG Marvel R — not in 2026 data
    42: ['YARIS HYBRID'],
    43: ['YARIS CROSS HSD', 'YARIS CROSS HEV', 'YARIS CROSS', 'YARIS CROSS HYB'],
    44: ['G6'],
    45: ['BYD SEALION 7'],
    46: ['ZEEKR X'],
    47: None,   # Hyundai Ioniq 5 — not in 2026 data
    48: None,   # BMW iX3 — not in 2026 data
    49: None,   # BYD Atto 3 — not in 2026 data (Atto 2 is, but different model)
    50: None,   # Honda Accord — not in 2026 data
    51: ['ENYAQ 85'],
    52: None,   # Geely Geometry C — not in 2026 data
    53: None,   # Peugeot E208 — not in 2026 data
    54: ['C10'],              # LeapMotor C10 BEV (shared entry with row 17)
    55: None,   # VW ID.4 — not in 2026 data
    56: None,   # MG 5 — not in 2026 data
    57: ['ZS'],
    58: None,   # Eveasy Limo — not in 2026 data
    59: None,   # Seres 3 — not in 2026 data
    60: ['CHR', 'CHR PHEV'],
    61: ['EX30 SM', 'EX30 TM'],
    62: ['ZEEKR 001'],
    63: ['ZEEKR 7X'],
}

NOTES = {
    2:  'All Wrangler variants combined',
    10: 'All EHS variants combined (EHS/EHS5/EHS7)',
    13: 'SEAL U + BYD SEAL U combined',
    14: 'All Panamera PHEV variants combined',
    17: 'C10 entry covers both PHEV and BEV (rows 17 & 54)',
    21: 'All Camry HEV variants combined',
    26: 'Corolla sedan HEV only (excl. Corolla Cross)',
    27: 'RAV4 Hybrid/HSD combined',
    28: 'MODEL Y (all variants)',
    29: 'X5 xDrive50e PHEV variants only',
    36: 'XC40 B4 — note: B4 is mild hybrid; full PHEV variant may differ',
    42: 'Yaris Hybrid (excl. Yaris Cross)',
    43: 'All Yaris Cross HEV variants combined',
    54: 'C10 entry covers both PHEV and BEV (rows 17 & 54)',
    60: 'CHR + CHR PHEV combined',
    61: 'EX30 SM + TM combined',
}

# ── Build merged dataframe ────────────────────────────────
rows = []
for _, tam_row in df_tam.iterrows():
    row_num = int(tam_row['#'])
    mapping = MAPPING.get(row_num)
    if mapping:
        m = get_months(mapping)
    else:
        m = NONE.copy()

    rows.append({
        '#':           row_num,
        'Vendor':      tam_row['Vendor'],
        'Model':       tam_row['Model'],
        'Type':        tam_row['Type'],
        'Carzone YTD': tam_row['2026 YTD (Jan–Apr)'],
        'Jan':         m['Jan'],
        'Feb':         m['Feb'],
        'Mar':         m['Mar'],
        'Apr':         m['Apr'],
        'Gov Total':   m['GovTotal'],
        'Matched To':  ', '.join(mapping) if mapping else '— no 2026 data',
        'Notes':       NOTES.get(row_num, ''),
    })

df_merged = pd.DataFrame(rows)

# ── Build Excel ───────────────────────────────────────────
DARK_BLUE   = PatternFill('solid', fgColor='1F4E79')
MID_BLUE    = PatternFill('solid', fgColor='2E75B6')
LIGHT_BLUE  = PatternFill('solid', fgColor='D6E4F0')
ORANGE      = PatternFill('solid', fgColor='ED7D31')
GREEN_FILL  = PatternFill('solid', fgColor='E2EFDA')
GRAY        = PatternFill('solid', fgColor='F2F2F2')
NO_DATA_FILL= PatternFill('solid', fgColor='FFF2CC')   # yellow — no match
WHITE_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BOLD        = Font(name='Arial', bold=True, size=10)
NORMAL      = Font(name='Arial', size=10)
GRAY_FONT   = Font(name='Arial', size=9, color='808080', italic=True)
NOTE_FONT   = Font(name='Arial', italic=True, color='595959', size=8)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left',   vertical='center')
NUM_FMT     = '#,##0'

wb = Workbook()
ws = wb.active
ws.title = 'Merged 2026 YTD'

headers = [
    ('#',            4),
    ('Vendor',      13),
    ('Model',       24),
    ('Type',         6),
    ('Carzone YTD', 13),   # from carzone scrape
    ('Jan',         10),
    ('Feb',         10),
    ('Mar',         10),
    ('Apr',         10),
    ('Gov Total',   12),   # from Transport Ministry
    ('Matched To',  32),
    ('Notes',       36),
]

for ci, (h, w) in enumerate(headers, 1):
    cell = ws.cell(1, ci, h)
    cell.font = WHITE_FONT
    cell.fill = MID_BLUE if ci in (6, 7, 8, 9) else DARK_BLUE
    cell.alignment = CENTER
    ws.column_dimensions[get_column_letter(ci)].width = w

for ri in range(2, len(df_merged) + 2):
    row = None  # unused, access via df_merged.iloc
    no_data = df_merged.iloc[ri - 2]['Matched To'] == '— no 2026 data'
    base_fill = NO_DATA_FILL if no_data else (GRAY if ri % 2 == 0 else PatternFill())

    r = df_merged.iloc[ri - 2]
    vals = [r['#'], r['Vendor'], r['Model'], r['Type'],
            r['Carzone YTD'],
            r['Jan'], r['Feb'], r['Mar'], r['Apr'],
            r['Gov Total'],
            r['Matched To'],
            r['Notes'],
    ]

    for ci, val in enumerate(vals, 1):
        cell = ws.cell(ri, ci, val if val is not None else '')
        cell.font = GRAY_FONT if no_data and ci in (6,7,8,9,10) else NORMAL
        cell.fill = base_fill
        cell.alignment = LEFT if ci in (2, 3, 11, 12) else CENTER
        if ci in (5, 6, 7, 8, 9, 10) and isinstance(val, (int, float)):
            cell.number_format = NUM_FMT

# Totals row
tr = len(df_merged) + 2
ws.cell(tr, 1, 'TOTAL').font = BOLD
ws.cell(tr, 1).fill = ORANGE
ws.cell(tr, 1).font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
for ci in (5, 6, 7, 8, 9, 10):
    col_l = get_column_letter(ci)
    c = ws.cell(tr, ci, '=SUM(' + col_l + '2:' + col_l + str(tr-1) + ')')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = ORANGE
    c.alignment = CENTER
    c.number_format = NUM_FMT

# Legend row
leg_row = tr + 2
ws.cell(leg_row, 1, 'Legend:').font = BOLD
ws.cell(leg_row+1, 1).fill = NO_DATA_FILL
ws.cell(leg_row+1, 2, 'Yellow = model not found in Jan-Apr 2026 government registration data (may be discontinued or 0 sales)').font = NOTE_FONT
ws.cell(leg_row+2, 2, 'Carzone YTD: scraped from carzone.co.il (data as of April 2026, published May 3 2026)').font = NOTE_FONT
ws.cell(leg_row+3, 2, 'Gov Total: Israeli Ministry of Transport via data.gov.il (Jan-Apr 2026, published May 3 2026)').font = NOTE_FONT

ws.freeze_panes = 'E2'
ws.auto_filter.ref = 'A1:L' + str(len(df_merged) + 1)

out_path = 'output/cars_tam_merged_2026.xlsx'
wb.save(out_path)
print('Saved:', out_path)
print('Rows:', len(df_merged))
matched = sum(1 for v in df_merged['Matched To'] if v != '— no 2026 data')
print('Models matched to gov data:', matched, '/', len(df_merged))
